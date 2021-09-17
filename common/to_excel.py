# -*- coding: utf-8 -*-

import json
import os
import uuid
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.FILE_PATH import DOWNLOAD_PATH

class CreateCases(object):
    """
    从rap2中导入数据
    """

    def __init__(self, name):
        #文件名拼接uuid
        self.file_name = f"{name}-{uuid.uuid4().hex}.xlsx"
        #文件路径
        self.file_path = os.path.join(DOWNLOAD_PATH, self.file_name)
        self.title = ['id','title','description','url','method','headerData','queryData','data','assertAction','actual','result','rearAction']
        # self.title = ['用例ID','用例标题','用例描述','接口路径','请求方式','请求头','查询参数','请求参数','断言操作','实际请求结果','测试结果','后置操作']
        #高度
        self.height = 25.0
        #宽度
        self.width = 25.0
        # 设置填充颜色
        self.fill = PatternFill('solid', fgColor='00ff00')
        # 设置字体
        self.font = Font(u'宋体', size=12)
        # 设置居中
        self.alignment = Alignment(horizontal='center', vertical='center')
        # 设置框线
        self.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

    def create_file(self, data):
        """
        :param data: 接口数据
        :return:
        """
        wb = Workbook()
        for index, dto in enumerate(data):
            #根据模块名生成各个sheet页
            wb.create_sheet(dto['modules'], index)
            ws = wb[dto['modules']]
            for column, title in enumerate(self.title):
                #设置列名
                ws_obj = ws.cell(1, column+1, title)
                ws_obj.font = self.font
                ws_obj.alignment = self.alignment
                ws_obj.fill = self.fill
                ws_obj.border = self.border
                #设置列宽
                ws.column_dimensions[get_column_letter(column+1)].width = self.width
            # 设置行高=25
            ws.row_dimensions[1].height = self.height
            #遍历case_data数据
            for i, data in enumerate(dto['cases_data']):
                for k,v in enumerate(data.keys()):
                    if isinstance(data[v],dict):
                        data[v] = json.dumps(data[v],ensure_ascii=False)
                    #根据行列写入数据
                    ws_obj = ws.cell(i+2, k+1, data[v])
                    ws_obj.font = self.font
                    ws_obj.alignment = self.alignment
                    ws.row_dimensions[i+2].height = self.height
        #openpyxl 默认会生成一个sheet页
        ws_sheet = wb["Sheet"]
        #强迫症将其删除
        wb.remove(ws_sheet)
        #保存文件
        wb.save(self.file_path)
        wb.close()
        return self.file_name

    def write_data(self, sheet_name, row, column, value):
        """
        :param sheet_name: 工作表
        :param row: 行
        :param column: 列
        :param value: 值
        :return:
        """
        wb = load_workbook(self.file_path)
        ws = wb[sheet_name]
        ws_obj = ws.cell(row, column, value)
        ws_obj.font = self.font
        ws_obj.alignment = self.alignment
        ws.row_dimensions[row].height = self.height
        wb.save(self.file_path)
        wb.close()



if __name__ == '__main__':
    create_excel = CreateCases('test')
    create_excel.create_file(407)
